# ( ) : 
import PySimpleGUI as sg
import os
from openpyxl import load_workbook


def findPath():
    temp = str(os.path.dirname(__file__) + "/path.txt")
    f = open(temp, "r")
    
    if len(f.read()) <= 0:
        layout = [
            [sg.Text("Vælg excel ark med opgaver")],
            [sg.Input(), sg.FileBrowse()],
            [sg.Submit()]
            ]
        window = sg.Window("Vælg fil").Layout(layout)

        values = window.Read()
        values = values[1]
        
        path1 = values[0]
        
        f = open(temp, "w")
        f.write(path1)
    else:
        f = open(temp, "r")
        path1 = f.read()
    
    path2 = path1[:-50]
    return path1, path2


def returnOpgaver(pathToOpgaveOversigt):
    wb = load_workbook(filename = pathToOpgaveOversigt)
    sheet_ranges = wb["Ark1"]
    kapDict = {}
    
    for row in range(sheet_ranges.max_row):
        kap = sheet_ranges["A" + str(row + 1)].value
        if kap is None:
            break
        
        value = sheet_ranges["B" + str(row + 1)].value
        value = value.split(",")
        # print(value)

        opgaver = []
        
        for opgave in value:
            opgave = opgave.split("-")
            if type(opgave) == list and len(opgave) == 2:
                for element in list(range(int(opgave[0]), int(opgave[1]) + 1)):
                    opgaver.append(str(element).zfill(3) + "a")
            else:
                opgaver.append(str(*opgave).zfill(3) + "a")

        kapDict[kap] = opgaver
    return kapDict

    
def getOwnFiles(path, kapitler):
    ownFiles = {}
    temp = os.listdir(path)
    for item in range(len(temp)):
        if temp[item] != "Mirsad Kadribasic - ZZMatInfoMappe" and temp[item][:6] == "Mirsad":
            path = "{}/{}".format(path, temp[item])
            kapList =  os.listdir(path)
            break
    
    for x in range(len(kapitler)):
        ownFiles[kapitler[x]] = []
        tempPath = "{}/{}".format(path, kapitler[x])
        opgaver = os.listdir(tempPath) 
        for y in range(len(opgaver)):
            if opgaver[y][-4:] == ".pdf" and opgaver[y][0] != ".":
                ownFiles[kapitler[x]].append(opgaver[y][:4])

    for i in ownFiles.items():
        print(i)
    return ownFiles

def compare(opgaver, egneOpgaver):
    manglende = {}
    m = []
    for x in opgaver.items():
        kap = x[0]
        opg = x[1]
        manglende[kap] = []
        for i in range(len(opg)):
            if opg[i] in egneOpgaver[kap]:
                continue
                #print(opg[i])
            else:
                manglende[kap].append(opg[i])
                m.append("{}: {}".format(kap, opg[i]))
    if len(m) <= 0:
        x = True
    else:
        x = False
    return manglende, x, m


def peopleInPrison(pathToOversigtPlus):
    wb = load_workbook(filename = pathToOversigtPlus)
    sheet_ranges = wb["Ark1"]
    prisonList = []
    for n in range (43, 74):
        name = sheet_ranges["A" + str(n)].value
        value = sheet_ranges["C" + str(n)].value
        if value != 100:
            prisonList.append(name)
    return prisonList


if __name__ == "__main__":
    path = findPath()
    opgaveFiles = returnOpgaver(path[0])
    kap = list(opgaveFiles.keys())
    listWithPeopleInPrison = peopleInPrison("C:\\Users\\nikol\\AARHUS TECH\\Mirsad Kadribasic - ZZMatInfoMappe\\Oversigt18RPLUS.xlsx")
    ownFiles = getOwnFiles(path[1], kap)
    manglende, x, m = compare(opgaveFiles, ownFiles)
    print(m)
